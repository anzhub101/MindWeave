from __future__ import annotations

import csv
import hashlib
import io
import re
from pathlib import Path

from docx import Document as DocxDocument
from fastapi import UploadFile
from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.config import get_settings
from app.models.runtime import DocumentRecord
from app.services.ocr_service import ChunkrOCRService
from app.services.storage_service import StorageService


class DocumentProcessor:
    def __init__(self, storage_service: StorageService | None = None, ocr_service: ChunkrOCRService | None = None) -> None:
        settings = get_settings()
        self.settings = settings
        self.storage_root = settings.storage_root
        self.sample_data_root = settings.sample_data_root
        self.storage_service = storage_service or StorageService(settings=settings, storage_root=self.storage_root)
        self.ocr_service = ocr_service or ChunkrOCRService(settings=settings)

    async def store_uploads(self, task_id: str, files: list[UploadFile]) -> list[DocumentRecord]:
        documents: list[DocumentRecord] = []
        for file in files:
            raw = await file.read()
            filename = file.filename or "document.bin"
            media_type = file.content_type or "application/octet-stream"
            documents.append(self._persist_file(task_id, filename, media_type, raw))
        return documents

    def load_sample_pack(self, task_id: str, pack_id: str = "invisium_fy2026") -> list[DocumentRecord]:
        pack_root = self.sample_data_root / pack_id
        documents: list[DocumentRecord] = []
        for path in sorted(pack_root.iterdir()):
            if path.is_file():
                documents.append(
                    self._persist_file(
                        task_id=task_id,
                        filename=path.name,
                        media_type=self._guess_media_type(path.suffix),
                        raw=path.read_bytes(),
                    )
                )
        return documents

    def _persist_file(self, task_id: str, filename: str, media_type: str, raw: bytes) -> DocumentRecord:
        safe_name = self._sanitize_filename(filename)
        relative_raw_path = f"{task_id}/{safe_name}"
        self.storage_service.storage_root = self.storage_root
        extracted_text, metadata = self._extract_text(safe_name, raw)
        raw_write = self.storage_service.store_bytes("uploads", relative_raw_path, raw, media_type)
        text_relative_path = f"{task_id}/{Path(safe_name).stem}.extracted.txt"
        text_write = self.storage_service.store_text("uploads", text_relative_path, extracted_text)

        sha256 = hashlib.sha256(raw).hexdigest()
        return DocumentRecord(
            id=f"doc_{sha256[:12]}",
            name=safe_name,
            media_type=media_type,
            storage_path=raw_write.locator,
            text_path=text_write.locator,
            sha256=sha256,
            extracted_text=extracted_text,
            metadata={
                **metadata,
                "storage_backend": self.storage_service.backend,
                "local_storage_path": str(raw_write.local_path),
                "local_text_path": str(text_write.local_path),
            },
        )

    def _extract_text(self, filename: str, raw: bytes) -> tuple[str, dict]:
        suffix = Path(filename).suffix.lower()
        if suffix == ".pdf":
            try:
                reader = PdfReader(io.BytesIO(raw))
                pages = [page.extract_text() or "" for page in reader.pages]
                text = "\n\n".join(pages).strip()
                metadata = {"page_count": len(pages), "ocr_used": False}
            except Exception as exc:
                text = ""
                metadata = {"page_count": 0, "ocr_used": False, "parser_error": str(exc)}
            if self._should_fallback_to_ocr(text):
                ocr_text, ocr_metadata = self.ocr_service.extract_text(raw, filename)
                if ocr_text.strip():
                    return ocr_text, {**metadata, **ocr_metadata, "ocr_used": True}
            return text, metadata
        if suffix == ".docx":
            document = DocxDocument(io.BytesIO(raw))
            paragraphs = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
            return "\n".join(paragraphs), {"paragraph_count": len(paragraphs)}
        if suffix == ".csv":
            return self._extract_csv(raw)
        if suffix == ".xlsx":
            return self._extract_xlsx(raw)

        decoded = raw.decode("utf-8", errors="ignore")
        return decoded, {"line_count": len(decoded.splitlines())}

    def _extract_csv(self, raw: bytes) -> tuple[str, dict]:
        decoded = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = [dict(row) for row in reader]
        lines = []
        for row in rows:
            values = [f"{key}: {value}" for key, value in row.items()]
            lines.append(", ".join(values))
        return "\n".join(lines), {"structured_rows": rows, "row_count": len(rows)}

    def _extract_xlsx(self, raw: bytes) -> tuple[str, dict]:
        workbook = load_workbook(io.BytesIO(raw), data_only=True)
        rows: list[dict[str, str]] = []
        lines: list[str] = []
        for sheet in workbook.worksheets:
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                continue
            headers = [str(value) if value is not None else "" for value in values[0]]
            for row_values in values[1:]:
                row = {
                    headers[index] or f"col_{index}": "" if value is None else str(value)
                    for index, value in enumerate(row_values)
                }
                rows.append(row)
                lines.append(", ".join(f"{key}: {value}" for key, value in row.items()))
        return "\n".join(lines), {"structured_rows": rows, "row_count": len(rows)}

    @staticmethod
    def _guess_media_type(suffix: str) -> str:
        return {
            ".txt": "text/plain",
            ".csv": "text/csv",
            ".pdf": "application/pdf",
            ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }.get(suffix.lower(), "application/octet-stream")

    def _should_fallback_to_ocr(self, text: str) -> bool:
        if not self.ocr_service.enabled():
            return False
        return len(text.strip()) < self.settings.chunkr_pdf_char_threshold

    @staticmethod
    def _sanitize_filename(filename: str) -> str:
        stem = re.sub(r"[^A-Za-z0-9._-]+", "_", filename).strip("._")
        return stem or "document"
